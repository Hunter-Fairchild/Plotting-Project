# Modules imported
"""This program uses Microsoft Excel which is one of the various applications provided
by Microsoft Office. """
import os
from typing import Union as typing_Union
import openpyxl as op
import openpyxl.chart as op_chart
import tkinter as tk
import math


# Window class
class Window:
    def __init__(self):
        # Window declaration and settings
        self.master = tk.Tk()
        self.master.title("Plotter")
        self.master.resizable(False, False)
        self.master.geometry("600x350")

        # tkinter frames
        self.left = tk.Frame(self.master)
        self.left.grid(column=0, row=0)
        self.right = tk.Frame(self.master)
        self.right.grid(column=1, row=0)
        self.bottom = tk.Frame(self.master)
        self.bottom.grid(column=0, columnspan=3, row=1, rowspan=2, sticky="W")

        # Correction label
        self.correction = tk.Label(self.bottom, text="", foreground="red")
        self.correction.grid(column=0, row=0, sticky="W")

        # Series name
        tk.Label(self.left, text="Series Name: ").grid(column=0, row=1, sticky="W")
        self.series_entry = tk.Entry(self.left, width=20)
        self.series_entry.grid(column=0, columnspan=2, row=2, sticky="W", padx=3)

        # Text Labels 
        tk.Label(self.left, text="Enter x-values: ").grid(column=0, row=3)
        tk.Label(self.left, text="Enter y-values: ").grid(column=1, row=3)

        # Text boxes for entering data
        self.x_text = tk.Text(self.left, width=9, height=15)
        self.x_text.grid(column=0, row=4, sticky="N")
        self.y_text = tk.Text(self.left, width=9, height=15)
        self.y_text.grid(column=1, row=4, sticky="N")

        # Check box for plotting regression line
        self.reg_var = tk.BooleanVar(self.master)
        self.reg_box = tk.Checkbutton(self.right, text="Plot Regression Line", variable=self.reg_var,
                                      onvalue=True, offvalue=False)
        self.reg_box.grid(column=0, row=0)

        # Button to make data and graph
        self.create_button = tk.Button(self.right, text="Create excel", command=self.command)
        self.create_button.grid(column=0, row=1)

        self.master.mainloop()

    # Command for create button
    def command(self):
        # Data stored from text widgets
        name = str(self.series_entry.get())
        xs = self.x_text.get("1.0", tk.END)
        xs = list(xs.split("\n")[:-1])
        ys = self.y_text.get("1.0", tk.END)
        ys = list(ys.split("\n")[:-1])
        self.create_sheet1(name, xs, ys)

    # Checks for valid inputs and formats the numbers
    def format_input(self, li: list, o_li: list) -> typing_Union[list, bool]:
        # Removes any blanks
        li = list(filter(lambda item: item != "" and item != " ", li))
        if len(li) == 0:
            self.correction.configure(text="Please enter values")
            return False
        # Checks for length match between the two lists
        # Accounts for the empty entries in the first li
        if len(li) != len(o_li)-o_li.count(""):
            self.correction.configure(text="Please enter same amount of values")
            return False

        # Checks for non-numbers
        try:
            temp = [int(item) for item in li]
        except ValueError:
            self.correction.configure(text="Please only enter only numbers.")
            return False

        # Puts all of the numbers into floats
        for num in range(0, len(li)):
            li[num] = float(li[num])
        return li

    # Function that creates the first excel sheet
    # Featuring the data and graph
    def create_sheet1(self, name: str, xs: list, ys: list):
        # Work book creation and storage of data
        wb = op.Workbook()
        data_sheet = wb["Sheet"]
        data_sheet.title = "Data and Graph"

        # Headers
        data_sheet["A1"] = "X-Values"
        data_sheet["B1"] = "Y-Values"

        # Checks values and ends function if values aren't correct
        if not self.format_input(xs, ys) or not self.format_input(ys, xs):
            return 
        xs, ys = self.format_input(xs, ys), self.format_input(ys, xs)
        self.correction.configure(text="")      # Resets error text

        # Adds values to the spreadsheet
        for x, y in zip(xs, ys):
            data_sheet.append([x, y])

        # Setting the x and y values into the chart
        xvalues = op_chart.Reference(data_sheet, min_col=1, max_col=1, min_row=2, max_row=len(xs)+1)
        yvalues = op_chart.Reference(data_sheet, min_col=2, max_col=2, min_row=2, max_row=len(ys)+1)
        if name == "":
            name = "No Name"
        series = op_chart.Series(yvalues, xvalues, title=name)
        series.marker.symbol = "circle"
        series.graphicalProperties.line.noFill = True

        # Chart creation
        chart = op_chart.ScatterChart()
        chart.title = "Data Graph"
        chart.x_axis.title = "X-Values"
        chart.y_axis.title = "Y-Values"
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None

        # Add series to chart and chart to sheet
        chart.series.append(series)
        if self.reg_var:
            chart.series.append(self.plot_regression(data_sheet, xs, ys))
        data_sheet.add_chart(chart, "D3")

        # Adds second excel sheet
        self.create_sheet2(wb, xs, ys)

        # Saves and opens the file automatically 
        wb.save("Example.xlsx")
        os.system("Example.xlsx")

    @staticmethod
    def plot_regression(data_sheet, xs: list, ys: list) -> op_chart.Series:
        """Function to plot the regression line. Gets the regression points than plots the first and last point in
        a series."""
        data_sheet.append(["  "])
        data_sheet.append(["Regression Line: ", Calculations().regression_line(xs, ys).format(x="x")])
        for num in range(0, len(ys)):
            ys[num] = round(eval(Calculations().regression_line(xs, ys).format(x=str(xs[num]))), 2)
        xy = sorted(tuple((x, y) for x, y in zip(xs, ys)), key=lambda t: t[0])
        data_sheet.append([xy[0][0], xy[0][1]])
        data_sheet.append([xy[-1][0], xy[-1][1]])
        xvalues = op_chart.Reference(data_sheet, min_col=1, max_col=1, min_row=len(xs)+4, max_row=len(xs)+5)
        yvalues = op_chart.Reference(data_sheet, min_col=2, max_col=2, min_row=len(ys)+4, max_row=len(ys)+5)
        return op_chart.Series(yvalues, xvalues, title="Regression Line: ")

    # Function that creates the second excel sheet
    # Featuring all the basic statistics
    @staticmethod
    def create_sheet2(workbook: op.Workbook, xs: list, ys: list):
        calc = Calculations()
        stats_sheet = workbook.create_sheet(title="Data Statistics")
        stats_sheet.append(["Maximum of x: ", calc.maximum(xs)])
        stats_sheet.append(["Maximum of y: ", calc.maximum(ys)])
        stats_sheet.append(["Minimum of x: ", calc.minimum(xs)])
        stats_sheet.append(["Minimum of y: ", calc.minimum(ys)])
        stats_sheet.append(["Sum of x: ", calc.summ(xs)])
        stats_sheet.append(["Sum of y: ", calc.summ(ys)])
        stats_sheet.append(["Median of x: ", calc.median(xs)])
        stats_sheet.append(["Median of y: ", calc.median(ys)])
        stats_sheet.append(["Count of x: ", calc.count(xs)])
        stats_sheet.append(["Count of y: ", calc.count(ys)])
        stats_sheet.append(["Mean of x: ", calc.mean(xs)])
        stats_sheet.append(["Mean of y: ", calc.mean(ys)])
        stats_sheet.append(["Range of x: ", calc.range(xs)])
        stats_sheet.append(["Range of y: ", calc.range(ys)])
        stats_sheet.append(["Standard Deviation of x: ", calc.standard_deviation(xs)])
        stats_sheet.append(["Standard Deviation of y: ", calc.standard_deviation(ys)])
        stats_sheet.append(["Correlation: ", calc.correlation(xs, ys)])
        stats_sheet.append(["Slope: ", calc.slope(xs, ys)])
        stats_sheet.append(["Y-Intercept: ", calc.y_intercept(xs, ys)])
        stats_sheet.append(["Regression Line: ", calc.regression_line(xs, ys)])


# Class that holds all functions to calculate statistics
class Calculations:
    @staticmethod
    def maximum(values: list) -> float:
        """Returns highest value in the list passed
        Maximum is set to first value of list and then loops through the list and if any value
        of the list is greater than the maximum, that value becomes the new maximum"""
        maximum = values[0]
        for value in values[1:]:
            if value > maximum:
                maximum = value
        return maximum

    @staticmethod
    def minimum(values: list) -> float:
        """Returns lowest value in the list passed
        minimum is set to first value of list and then loops through the list and if any value
        of the list is less than the minimum, that value becomes the new minimum"""
        minimum = values[0]
        for value in values[1:]:
            if value < minimum:
                minimum = value
        return minimum

    @staticmethod
    def summ(values: list) -> float:
        """Returns sum of all values in the list passed
        Loops through the list adding every value of the list to the total"""
        total = 0
        for value in values:
            total += value
        return total

    @staticmethod
    def median(values: list) -> float:
        """Returns the median (middle value) of the list passed.
        Sorts the list from least to greatest, then decides if the length is even or odd.
        If length is odd it returns the middle value of the sorted list.
        If length is even it returns the sum of the two middle most values in the sorted list divided by two."""
        values = sorted(values)
        if len(values) % 2 == 0:
            middle_values = values[int(len(values)/2)-1] + values[int(len(values)/2)]
            return middle_values/2
        return values[int(len(values)/2)]

    @staticmethod
    def count(values: list) -> float:
        """Returns the length of the list passed."""
        return len(values)

    def mean(self, values: list) -> float:
        """Returns average of all values in the list passed
        Loops through the list to get the total than divides the total by the length of the list"""
        total = self.summ(values)
        return total / len(values)

    def range(self, values: list) -> float:
        """Returns the range between the highest and lowest value in the list passed.
        Finds the maximum and minimum value in the list and returns the maximum minus the minimum."""
        maximum = self.maximum(values)
        minimum = self.minimum(values)
        return maximum - minimum

    def standard_deviation(self, values: list) -> float:
        """Returns the standard deviation of the values passed in the list.
        Finds the mean of the values, then loops through the list adding the value minus the mean squared to the
        standard deviation. The standard deviation is then divided by the count minus one and returns the square root
        of the standard deviation. """
        if self.count(values) <= 1:
            return 0
        sd = 0
        mean = self.mean(values)
        for value in values:
            sd += (value-mean)**2
        sd /= (self.count(values)-1)
        return math.sqrt(sd)

    def correlation(self, xs: list, ys: list) -> float:
        """Returns correlation of the x values and y values passed (Note: 'r' is the variable for the correlation).
        Finds the mean and standard deviation of both x and y. Loops through both x and y list and adds the (instance of
        x - mean of x) divided by the standard deviation of x times the (instance of y - mean of y) divided by the
        standard deviation of y to the correlation. Then, take the correlation and times it by one divided by (the
        amount of either x or y values and subtract one from the amount). The function then returns the correlation."""
        if self.count(xs) <= 1:
            return 0
        r = 0
        mean_x = self.mean(xs)
        mean_y = self.mean(ys)
        sx = self.standard_deviation(xs)
        sy = self.standard_deviation(ys)
        for x, y in zip(xs, ys):
            r += ((x-mean_x)/sx)*((y-mean_y)/sy)
        r *= (1/(self.count(xs)-1))
        return r

    def slope(self, xs: list, ys: list) -> float:
        """Returns the slope of the slope of the x values and y values passed.
        Finds the correlation and standard deviations of x and y. Multiples the correlation by the standard deviation
        of y divided by the standard deviation of x"""
        if self.count(xs) <= 1:
            return 0
        r = self.correlation(xs, ys)
        s_x = self.standard_deviation(xs)
        s_y = self.standard_deviation(ys)
        return r*s_y/s_x

    def y_intercept(self, xs: list, ys: list) -> float:
        """Returns the y-intercept of the x values and y values passed.
        Finds the mean of x, mean of y and slope of x and y, then returns the mean of y minus (the mean of x times the
        slope)"""
        if self.count(xs) <= 1:
            return 0
        mean_y = self.mean(ys)
        mean_x = self.mean(xs)
        slope = self.slope(xs, ys)
        return mean_y - (mean_x*slope)

    def regression_line(self, xs: list, ys: list) -> str:
        """Returns the regression line of the x values and y values passed.
        Finds the slope and y-intercept of x and y. Returns a string containing the regression line formula (which is
        y=slope*x + y-intercept) for the x values and y values entered. The string that is returned may be formatted so
        that it can be used as both a written formula and an actual math formula"""
        slope = round(self.slope(xs, ys), 2)
        y_intercept = round(self.y_intercept(xs, ys), 2)
        return f"{slope}*" + "{x} + " + f"{y_intercept}"


Window()
